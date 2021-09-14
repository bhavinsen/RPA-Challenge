import os
import time
import openpyxl
from selenium import webdriver
from bs4 import BeautifulSoup as bs
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager


def check_workbook(wb_path):
    if os.path.exists(wb_path):
        return openpyxl.load_workbook(wb_path)
    return openpyxl.Workbook()


wb = check_workbook(r"./Excel/Agencies.xlsx")
if "Agencies" not in wb.sheetnames:
    sheet = wb["Sheet"]
    sheet.title = "Agencies"
else:
    sheet = wb["Agencies"]

sheet.cell(row=1, column=1).value = "Agency Name"
sheet.cell(row=1, column=2).value = "Amount"


def chrome_driver():
    """
    initialize chrome driver with headless
    :return:
    """
    chrome_options = Options()
    chrome_options.add_argument("no-sandbox")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--disable-dev-shm-ussage")
    chrome_options.add_argument("--allow-running-insecure-content")
    chrome_options.add_argument("--disable-web-security")
    prefs = {
        "profile.default_content_settings.popups": 0,
        "download.default_directory": os.getcwd() + "/Download/",
        "directory_upgrade": True,
    }
    chrome_options.add_experimental_option("prefs", prefs)
    # chrome_options.add_argument("--headless")
    return webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)


driver = chrome_driver()
driver.maximize_window()

driver.get("https://itdashboard.gov/")

download_links = []
individual_urls = []

driver.find_element_by_xpath('//a[contains(text(), "DIVE IN")]').click()
time.sleep(5)
agency_soup = bs(driver.page_source, "lxml")

for index, agency in enumerate(
    agency_soup.find("div", {"id": "agency-tiles-widget"}).find_all("div", {"class": "col-sm-12"}),
    start=1,
):

    details = agency.find_all("a")[1]

    agency_name = details.find_all("span")[0].text.strip()
    amount = details.find_all("span")[1].text.strip()
    sheet.cell(row=index + 1, column=1).value = agency_name
    sheet.cell(row=index + 1, column=2).value = amount
    individual_urls.append(f"https://itdashboard.gov{details['href']}")


if "Individual Investments" not in wb.sheetnames:
    wb.create_sheet("Individual Investments")
    sheet = wb["Individual Investments"]
else:
    sheet = wb["Individual Investments"]

INDIVIDUAL_HEADERS = [
    "UII",
    "Bureau",
    "Investment Title",
    "Total FY2021 Spending ($M)",
    "Type",
    "CIO Rating",
    "# of Projects",
]

for index, header in enumerate(INDIVIDUAL_HEADERS, start=1):
    sheet.cell(row=1, column=index).value = header
    
last_row = 2
for individual in individual_urls:
    driver.get(individual)
    time.sleep(10)
    Select(driver.find_element_by_name("investments-table-object_length")).select_by_value("-1")
    time.sleep(20)

    individual_soup = bs(driver.page_source, "lxml")

    for tr in (
        individual_soup.find("table", {"id": "investments-table-object"})
        .find("tbody")
        .find_all("tr")
    ):
        td = tr.find_all("td")
        UII = td[0].text.strip()
        bureau_name = td[1].text.strip()
        investment_title = td[2].text.strip()
        total_cyspending = td[3].text.strip()
        investment_type = td[4].text.strip()
        cio_rating = td[5].text.strip()
        number_of_projects = td[6].text.strip()
        
        sheet.cell(row=last_row, column=1).value = UII
        sheet.cell(row=last_row, column=2).value = bureau_name
        sheet.cell(row=last_row, column=3).value = investment_title
        sheet.cell(row=last_row, column=4).value = total_cyspending
        sheet.cell(row=last_row, column=5).value = investment_type
        sheet.cell(row=last_row, column=6).value = cio_rating
        sheet.cell(row=last_row, column=7).value = number_of_projects
        
        if td[0].find("a"):
            download_links.append(f"https://itdashboard.gov{td[0].find('a')['href']}")
        last_row += 1
        
wb.save(r"./Excel/Agencies.xlsx")

for download in download_links:
    driver.get(download)
    time.sleep(5)
    driver.find_element_by_xpath('//*[@id="business-case-pdf"]/a').click()
    time.sleep(3)
