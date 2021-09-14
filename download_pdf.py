import os
import openpyxl
import urllib.request
from requests import Session


# open workbook if existing otherwise create new one
def check_workbook(wb_path):
    if os.path.exists(wb_path):
        return openpyxl.load_workbook(wb_path)
    return openpyxl.Workbook()


wb = check_workbook("./Excel/Agencies.xlsx")
if "Agencies" not in wb.sheetnames:
    sheet = wb["Sheet"]
    sheet.title = "Agencies"
else:
    sheet = wb["Agencies"]

sheet.cell(row=1, column=1).value = "Agency Name"
sheet.cell(row=1, column=2).value = "Amount"


BASE_URL = "https://itdashboard.gov/"
querystring = {
    "draw": " 2",
    "columns[0][data]": " 0",
    "columns[0][name]": " ",
    "columns[0][searchable]": " true",
    "columns[0][orderable]": " true",
    "columns[0][search][value]": " ",
    "columns[0][search][regex]": " false",
    "columns[1][data]": " 1",
    "columns[1][name]": " ",
    "columns[1][searchable]": " true",
    "columns[1][orderable]": " true",
    "columns[1][search][value]": " ",
    "columns[1][search][regex]": " false",
    "columns[2][data]": " 2",
    "columns[2][name]": " ",
    "columns[2][searchable]": " true",
    "columns[2][orderable]": " true",
    "columns[2][search][value]": " ",
    "columns[2][search][regex]": " false",
    "columns[3][data]": " 3",
    "columns[3][name]": " ",
    "columns[3][searchable]": " true",
    "columns[3][orderable]": " true",
    "columns[3][search][value]": " ",
    "columns[3][search][regex]": " false",
    "columns[4][data]": " 4",
    "columns[4][name]": " ",
    "columns[4][searchable]": " true",
    "columns[4][orderable]": " true",
    "columns[4][search][value]": " ",
    "columns[4][search][regex]": " false",
    "columns[5][data]": " 5",
    "columns[5][name]": " ",
    "columns[5][searchable]": " true",
    "columns[5][orderable]": " true",
    "columns[5][search][value]": " ",
    "columns[5][search][regex]": " false",
    "columns[6][data]": " 6",
    "columns[6][name]": " ",
    "columns[6][searchable]": " true",
    "columns[6][orderable]": " true",
    "columns[6][search][value]": " ",
    "columns[6][search][regex]": " false",
    "columns[7][data]": " 7",
    "columns[7][name]": " ",
    "columns[7][searchable]": " true",
    "columns[7][orderable]": " true",
    "columns[7][search][value]": " ",
    "columns[7][search][regex]": " false",
    "columns[8][data]": " 8",
    "columns[8][name]": " ",
    "columns[8][searchable]": " true",
    "columns[8][orderable]": " true",
    "columns[8][search][value]": " ",
    "columns[8][search][regex]": " false",
    "order[0][column]": " 5",
    "order[0][dir]": " asc",
    "order[1][column]": " 0",
    "order[1][dir]": " asc",
    "start": " 0",
    "length": " -1",
    "search[value]": " ",
    "search[regex]": " false",
    "full": " 1",
}


def convert_to_currency(val):
    if val > 00:
        if val >= 10000:
            return "$" + str(round((val / 1000), 0)) + "B"
        elif val >= 1000:
            return "$" + str(round((val / 1000), 1)) + "B"
        elif val >= 100:
            return "$" + str(round(val, 0)) + "M"
        elif val >= 1:
            return "$" + str(round(val, 0)) + "M"
        else:
            return "$" + (round(val * 100, 0)) + "K"
    elif not val:
        return "--"
    return "$0"


# get cookies from website
session = Session()
session.get("https://itdashboard.gov/")


cookie = session.cookies.get_dict()
headers = {
    "Cookie": "; ".join([f"{key}={value}" for key, value in cookie.items()]),
    "Host": "itdashboard.gov",
    "Referer": "https://itdashboard.gov/",
    "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36",
}


def download_file(download_url, filename):
    request = urllib.request.Request(download_url)
    request.add_header("Cookie", headers["Cookie"])
    request.add_header("Host", headers["Host"])
    request.add_header("Referer", headers["Referer"])
    request.add_header("User-Agent", headers["User-Agent"])
    response = urllib.request.urlopen(request)

    file = open(f"./Download/{filename}.pdf", "wb")

    file.write(response.read())

    file.close()


agencies = session.get(
    "https://itdashboard.gov/api/v1/ITDB2/visualization/govwide/agencyTiles", headers=headers
).json()
last_row = 1
individual_urls = []
for index, agency in enumerate(agencies["result"], start=1):

    sheet.cell(row=index + 1, column=1).value = agency["agencyName"]
    sheet.cell(row=index + 1, column=2).value = convert_to_currency(agency["totalSpendingCY"])
    individual_urls.append(
        f"{BASE_URL}api/v1/ITDB2/visualization/agency/investmentsTable/agencyCode/{agency['agencyCode']}"
    )

if "Individual Investments" not in wb.sheetnames:
    wb.create_sheet("Individual Investments")
    sheet = wb["Individual Investments"]
else:
    sheet = wb["Individual Investments"]

INDIVIDUAL_HEADERS = [
    "UII",
    "Bureau",
    "Investment Title",
    "Total FY2021 Spending ($M)",
    "Type",
    "CIO Rating",
    "# of Projects",
]

for index, header in enumerate(INDIVIDUAL_HEADERS, start=1):
    sheet.cell(row=1, column=index).value = header

last_row = 2
for url in individual_urls:
    response = session.get(url=url, headers=headers, params=querystring).json()

    for record in response["result"]:
        sheet.cell(row=last_row, column=1).value = record["UII"]
        sheet.cell(row=last_row, column=2).value = record["bureauName"]
        sheet.cell(row=last_row, column=3).value = record["investmentTitle"]
        sheet.cell(row=last_row, column=4).value = f"${record['totalCySpending']}"
        sheet.cell(row=last_row, column=5).value = record["investmentType"]
        sheet.cell(row=last_row, column=6).value = record["cioRating"]
        sheet.cell(row=last_row, column=7).value = record["numberOfProjects"]
        if record["cioRating"]:
            download_url = f"https://itdashboard.gov/api/v1/ITDB2/businesscase/pdf/generate/uii/{record['UII']}"
            try:
                download_file(download_url, record["UII"])
            except:
                pass
        last_row += 1

wb.save(r"./Excel/Agencies.xlsx")
